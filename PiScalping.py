import datetime
import http.client
import os
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import requests
from curl_cffi import requests as req
import traceback
import time
import httpx
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import undetected_chromedriver as uc
from playwright.sync_api import sync_playwright
import random
from pyvirtualdisplay import Display
from fake_useragent import UserAgent

ua = UserAgent()
headers = {'User-Agent': ua.random}

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate',
    'Connection': 'keep-alive',
    'Referer': 'https://www.google.com/',
}
SENDER_EMAIL = "price.monitor.email@gmail.com"
RECEIVER_EMAIL = "price.monitor.email@gmail.com"
PASSWORD = "iejtkdfhxitktpgp"
driver_path = "/usr/bin/chromedriver"
EXCEL_FILE = "/usr/pi/Tick_Num.xlsx"
CELL = 'A1'

# TENNIS
URL_AO_INFO = 'https://ausopen.com/ticket-info#faqs'
URL_AO_TMAST = 'https://www.ticketmaster.com.au/browse/tennis-catid-31/sports-rid-10004?startDate=2025-01-12&endDate=2025-01-26'
URL_WIBLDN = 'https://www.wimbledon.com/en_GB/tickets/index.html'
URL_BARCA = 'https://www.barcelonaopenbancsabadell.com/en/news/' # Tickets URL = 'https://www.barcelonaopenbancsabadell.com/en/tickets-3/'
URL_CINCH = 'https://www.lta.org.uk/support-centre/major-events/cinch-championships/2025-london-championships/when-will-tickets-go-on-sale-for-the-2025-london-championship-and-how-can-i-purchase-them/'
US_OPEN = 'https://www.usopen.org/en_US/cms/feeds/tickets/individual_tickets.xml'
CINCINNATI = 'https://cincinnatiopen.com/tickets/single-sessions-2/'

URL_TAYLOR_OFFICIAL = 'https://www.taylorswift.com/tour/'
URL_COLDPLAY_OFFICIAL = 'https://www.coldplay.com/tour/'
URL_ADELE_OFFICIAL = 'https://www.adele.com/'
URL_DUA = 'https://www.dualipa.com/tour/'

URL_TAYLOR_TMAST = 'https://www.ticketmaster.com/taylor-swift-tickets/artist/1094215'
URL_ADELE_TMAST = 'https://www.ticketmaster.com/adele-tickets/artist/1159272'
URL_COLDPLAY_TMAST = 'https://www.ticketmaster.com/coldplay-tickets/artist/806431'
URL_DUA_TMAST = 'https://www.ticketmaster.com/dua-lipa-tickets/artist/2179476'
URL_CARPENTER_TMAST = 'https://www.ticketmaster.com/sabrina-carpenter-tickets/artist/2001092'
URL_RODRIGO_TMAST = 'https://www.ticketmaster.com/olivia-rodrigo-tickets/artist/2836194'
URL_EILISH_TMST = 'https://www.ticketmaster.com/billie-eilish-tickets/artist/2257710'
URL_BEYONCE_TMST = 'https://www.ticketmaster.com/beyonce-tickets/artist/894191'
URL_CABELLO_TMST = 'https://www.ticketmaster.com/camila-cabello-tickets/artist/2362959'
URL_METALLICA_TMST = 'https://www.ticketmaster.com/metallica-tickets/artist/735647'
URL_UEFA = ('https://uclf.hospitality.uefa.com/', 'https://uelf.hospitality.uefa.com/')
URL_FINAL4 = 'https://www.f4tickets.com/'


def send_email(subject, content, link):
    """Send an email notification."""
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = SENDER_EMAIL
    msg['To'] = RECEIVER_EMAIL

    html = f"""\
    <html>
      <body>
        <p><b>{content}</b></p>
        <p><b>Link:</b> <a href="{link}">{link}</a></p>
      </body>
    </html>
    """

    msg.attach(MIMEText(html, 'html'))

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(SENDER_EMAIL, PASSWORD)
        smtp.send_message(msg)


def save_tick_num_to_excel(tick_num, sheet_name):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    ws[CELL] = tick_num
    wb.save(EXCEL_FILE)


def load_tick_num_from_excel(sheet_name):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            return ws[CELL].value
    return None


def human_delay():
    time.sleep(random.uniform(2, 10))


def check_taylor_official_site(driver):
    print("\n")
    try:
        driver.get(URL_TAYLOR_OFFICIAL)
        soup = BeautifulSoup(driver.page_source, "lxml")
        # tour = soup.find('h2', class_='block__title block__title-tour').text.strip()
        no_tour = soup.find("div", class_="view-content view-tour")
        container = soup.find('div', class_='tour-grid--container')

        # if container:
        #     elements_count = len(container.find_all('div', class_='tour-grid--item'))
        #     print(f'The grid has {elements_count} elements.')
        #     date = container.find_all('div', class_='tour-grid--item')[elements_count - 1].text.strip()

        # if "The Eras Tour" in tour and "Sun, Dec 08, 2024" in date:
        if no_tour is not None and no_tour.get_text(strip=True):
            send_email("Tickets", "Taylor Swift Official", URL_TAYLOR_OFFICIAL)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Taylor Official")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Taylor Official")

    except Exception as e:
        send_email("Tickets for Taylor Official", e, URL_TAYLOR_OFFICIAL)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Taylor Official or site changed")
        print("Error in check_official_site:", e)


def check_ao_ticketmaster():
    print("\n")
    try:
        attempt = 0
        for _ in range(10):
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request("GET", f"/v2/general?url={URL_AO_TMAST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true")
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            shows = soup.find('p', class_='sc-1hz4ufn-1 hxUpt')
            if shows and shows.text.strip():
                shows = shows.text.strip()
                break
            else:
                attempt += 1
                print("attempt No ", attempt)
                print("No products found. Lets retry!\n")

        if attempt == 10:
            print("After 5 attempts no products found!\n")
            send_email("Tickets AO Ticketmaster 5 attempts", "5 attempts", URL_AO_TMAST)
            return
        if "Sorry... there are currently no upcoming events." in shows:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for AO Ticketmaster")
        else:
            send_email("Tickets", "Game is on", URL_AO_TMAST)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for AO Ticketmaster")
    except Exception as e:
        send_email("Tickets for AO Ticketmaster", str(e), URL_AO_TMAST)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for AO Ticketmaster or site changed")
        print("Error in check_ticketmaster:", e)


def check_ao_official_site():
    print("\n")
    try:
        response = requests.get(URL_AO_INFO, headers=HEADERS)
        soup = BeautifulSoup(response.content.decode('utf-8', 'ignore'), 'html.parser')
        shows = soup.find('span', class_='s1').text.strip()
        if "2024" in shows:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for AO Official")
        else:
            send_email("Tickets", "Game is on", URL_AO_INFO)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for AO Official")
    except Exception as e:
        send_email("Tickets for AO Official", e, URL_AO_INFO)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for AO Official or site changed")
        print("Error in check_official_site:", e)


def check_wimbledon(driver):
    print("\n")
    try:
        driver.get(URL_WIBLDN)
        soup = BeautifulSoup(driver.page_source, "lxml")
        shows = soup.find('div', class_='two-col textWrapContent margin clear-two clear-four').text.strip()

        if "The Wimbledon Public Ballot will open in September. Sign up to myWimbledon to be one of the first to hear about tickets to The Championships 2025." in shows:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Wimbledon")
        else:
            send_email("Tickets", "Game is on", URL_WIBLDN)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Wimbledon")
    except Exception as e:
        send_email("Tickets for Wimbledon", e, URL_WIBLDN)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Wimbledon", e)


def check_adele_official_site(driver):
    print("\n")
    try:
        driver.set_page_load_timeout(200)
        driver.get(URL_ADELE_OFFICIAL)
        time.sleep(5)
        driver.save_screenshot("ADELE_OF.png")
        soup = BeautifulSoup(driver.page_source, "lxml")
        tour = soup.find('div', class_='logowrap').text.strip()

        if "Tickets on sale now at Ticketmaster & Eventim" in tour:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Adele Official")
        else:
            send_email("Tickets", "Adele Official", URL_ADELE_OFFICIAL)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Adele Official")
    except Exception as e:
        send_email("Tickets for Adele Official", e, URL_ADELE_OFFICIAL)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Adele Official\n", e)


def check_coldplay_official_site():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 5
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request("GET",
                         f"/v2/general?url={URL_COLDPLAY_OFFICIAL}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true")
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            container = soup.find('div', class_='tour-listing__grid grid-outer')
            # print(container)
            if container:
                elements_count = len(container.find_all('div', class_='tour-card js-tour-card grid-inner'))
                print(f'The grid has {elements_count} elements.')
                date = container.find_all('div', class_='tour-card js-tour-card grid-inner')[
                    elements_count - 1].text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Coldplay official 5 attempts", "5 attempts", URL_COLDPLAY_OFFICIAL)
            return

        if "September 8 2025" in date:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Coldplay Official")
        else:
            send_email("Tickets", "Coldplay Official", URL_COLDPLAY_OFFICIAL)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Coldplay Official")
    except Exception as e:
        send_email("Tickets for Coldplay Official", e, URL_COLDPLAY_OFFICIAL)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              "Tickets are now available for Coldplay Official or site changed")
        print("Error in Coldplay Official:", e)


def check_dua_of():
    print("\n")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 720},
                device_scale_factor=1
            )

            # Open a new page
            page = context.new_page()

            # Visit the Dua Lipa Official page
            page.goto(URL_DUA)

            # Wait for the page to fully load (adjust if necessary)
            human_delay()  # Wait to mimic human behavior

            # Scroll down to simulate a human viewing the page
            page.mouse.wheel(0, 1000)
            human_delay()

            # Extract page content
            content = page.content()

            soup = BeautifulSoup(content, 'html.parser')
            # print(soup.prettify())

            container = soup.find('div', class_='container__splash__content')

            if container:
                elements_count = len(container.find_all('p'))
                print(f'There are {elements_count} shows.')
                date = container.find_all('p')[elements_count - 1].text.strip()
                print(date)

            if "16 october 2025" in date:
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Dua Lipa")
            else:
                send_email("Tickets for Dua Lipa", "Game is on", URL_DUA)
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Dua Lipa")

            browser.close()

    except Exception as e:
        # send_email("Tickets for Dua Lipa", e , URL_DUA)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              "Tickets are now available for Dua Lipa or site changed")
        print("Error in check_official_site:", e)


def check_taylor_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request("GET", f"/v2/general?url={URL_TAYLOR_TMAST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true")
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Taylor Ticketmaster 5 attempts", "5 attempts", URL_TAYLOR_TMAST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Taylor')
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Taylor Swift Ticketmaster", URL_TAYLOR_TMAST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Taylor')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Taylor Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Taylor')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Taylor Ticketmaster, Shows reduced")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Taylor Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_TAYLOR_TMAST)
        print(f"Error in check_Taylor_ticketmaster: {e}\nURL: {URL_TAYLOR_TMAST}")


def check_coldplay_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_COLDPLAY_TMAST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Coldplay Ticketmaster 5 attempts", "5 attempts", URL_COLDPLAY_TMAST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Coldplay')
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Coldplay Ticketmaster", URL_COLDPLAY_TMAST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Coldplay')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Coldplay Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Coldplay')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Coldplay Ticketmaster, Shows reduced")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Coldplay Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_COLDPLAY_TMAST)
        print(f"Error in check_Coldplay_ticketmaster: {e}\nURL: {URL_COLDPLAY_TMAST}")


def check_adele_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_ADELE_TMAST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Adele Ticketmaster 5 attempts", "5 attempts", URL_ADELE_TMAST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Adele')
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Adele Ticketmaster", URL_ADELE_TMAST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Adele')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Adele Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Adele')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Adele Ticketmaster, Shows reduced")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Adele Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_ADELE_TMAST)
        print(f"Error in check_Adele_ticketmaster: {e}\nURL: {URL_ADELE_TMAST}")


def check_dua_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_DUA_TMAST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Dua Lipa Ticketmaster 5 attempts", "5 attempts", URL_DUA_TMAST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Dua')
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Dua Lipa Ticketmaster", URL_DUA_TMAST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Dua')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Dua Lipa Ticketmaster")

        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Dua')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Dua Lipa Ticketmaster, Shows reduced")

        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Dua Lipa Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_DUA_TMAST)
        print(f"Error in check_Adele_ticketmaster: {e}\nURL: {URL_DUA_TMAST}")


def check_carpenter_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_CARPENTER_TMAST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Carpenter Ticketmaster 5 attempts", "5 attempts", URL_CARPENTER_TMAST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Carpenter') or 0
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Sabrina Carpenter Ticketmaster", URL_CARPENTER_TMAST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Carpenter')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Sabrina Carpenter Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Carpenter')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Sabrina Carpenter Ticketmaster, Shows reduced")

        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Carpenter Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_CARPENTER_TMAST)
        print(f"Error in check_Carpenter_ticketmaster: {e}\nURL: {URL_CARPENTER_TMAST}")


def check_rodrigo_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_RODRIGO_TMAST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Rodrigo Ticketmaster 5 attempts", "5 attempts", URL_RODRIGO_TMAST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Rodrigo') or 0
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Olivia Rodrigo Ticketmaster", URL_RODRIGO_TMAST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Rodrigo')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Olivia Rodrigo Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Rodrigo')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Olivia Rodrigo Ticketmaster, Shows reduced")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Olivia Rodrigo Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_RODRIGO_TMAST)
        print(f"Error in check_Rodrigo_ticketmaster: {e}\nURL: {URL_RODRIGO_TMAST}")


def check_eilish_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_EILISH_TMST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Eilish Ticketmaster 5 attempts", "5 attempts", URL_EILISH_TMST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Eilish') or 0
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Billie Eilish Ticketmaster", URL_EILISH_TMST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Eilish')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Billie Eilish Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Eilish')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Billie Eilish Ticketmaster, Shows reduced")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Billie Eilish Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_EILISH_TMST)
        print(f"Error in check_Eilish_ticketmaster: {e}\nURL: {URL_EILISH_TMST}")


def check_beyonce_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_BEYONCE_TMST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Beyonce Ticketmaster 5 attempts", "5 attempts", URL_BEYONCE_TMST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Beyonce') or 0
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Beyonce Ticketmaster", URL_BEYONCE_TMST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Beyonce')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Beyonce Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Beyonce')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Beyonce Ticketmaster, Shows reduced")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Beyonce Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_BEYONCE_TMST)
        print(f"Error in check_Beyonce_ticketmaster: {e}\nURL: {URL_BEYONCE_TMST}")


def check_cabello_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_CABELLO_TMST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Cabello Ticketmaster 5 attempts", "5 attempts", URL_CABELLO_TMST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Cabello') or 0
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Camila Cabello Ticketmaster", URL_CABELLO_TMST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Cabello')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Camila Cabello Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Cabello')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Camila Cabello Ticketmaster, Shows reduced")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Camila Cabello Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_CABELLO_TMST)
        print(f"Error in check_Cabello_ticketmaster: {e}\nURL: {URL_CABELLO_TMST}")


def check_metallica_ticketmaster():
    print("\n")
    try:
        for attempt in range(1, 6):  # attempt will range from 1 to 10
            # Scraping Ant
            conn = http.client.HTTPSConnection("api.scrapingant.com")
            conn.request(
                "GET",
                f"/v2/general?url={URL_METALLICA_TMST}&x-api-key=38329e714321477bba68cf51da3cc822&return_page_source=true"
            )
            res = conn.getresponse()
            datas = res.read()
            datas = datas.decode("utf-8")
            soup = BeautifulSoup(datas, 'html.parser')
            # print(soup.prettify())

            # Attempt to extract the number of shows
            shows_element = soup.find('span', class_='sc-8d839fd8-5 iMxMG')
            if shows_element and shows_element.text.strip():
                shows = shows_element.text.strip()
                break
            else:
                print(f"Attempt {attempt}: No products found. Let's retry!\n")

        else:  # This block executes if no break occurred in the loop (all attempts failed)
            print("After 5 attempts, no products found!\n")
            send_email("Tickets Metallica Ticketmaster 5 attempts", "5 attempts", URL_METALLICA_TMST)
            return

        # Parse the number of shows found
        Scraped_Tick_Num = int(shows.replace("Results", "").strip())
        Tick_Num = load_tick_num_from_excel(sheet_name='Metallica') or 0
        print("Scraped, Loaded:", Scraped_Tick_Num, Tick_Num)

        # Decision-making based on the number of shows
        if Scraped_Tick_Num > Tick_Num:
            send_email("Tickets Available", "Metallica Ticketmaster", URL_METALLICA_TMST)
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Metallica')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for Metallica Ticketmaster")
        elif Scraped_Tick_Num < Tick_Num:
            save_tick_num_to_excel(Scraped_Tick_Num, sheet_name='Metallica')
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Metallica Ticketmaster, Shows reduced")
        else:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "No available tickets for Metallica Ticketmaster, Shows number didn't change")

    except Exception as e:
        send_email("Tickets Error", str(e), URL_METALLICA_TMST)
        print(f"Error in check_Metallica_ticketmaster: {e}\nURL: {URL_METALLICA_TMST}")


def check_final4():
    print("\n")
    try:
        response = requests.get(URL_FINAL4, headers=HEADERS)
        soup = BeautifulSoup(response.content.decode('utf-8', 'ignore'), 'html.parser')
        # print(soup.prettify())

        show = soup.find('div', class_='pagebuilder-section-wrapper').text.strip()
        print(show)
        if "2024 Turkish Airlines EuroLeague Final Four Berlin" in show:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Final 4 2025")
        else:
            send_email("Tickets for Final 4 2025", "Game is on", URL_FINAL4)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Final 4 2025")
    except Exception as e:
        send_email("Tickets for Final 4 2025", e, URL_FINAL4)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Final 4 2025 or site changed")
        print("Error in check_official_site:", e)


def check_barcelona():
    # Tickets URL = 'https://www.barcelonaopenbancsabadell.com/en/tickets-3/'
    print("\n")
    try:
        response = requests.get(URL_BARCA, headers=HEADERS)
        soup = BeautifulSoup(response.content, 'html.parser')
        # print(soup.prettify())

        show = soup.find('div', class_='vc_grid-item vc_clearfix vc_col-sm-6 vc_grid-item-zone-c-right').text.strip()
        print(show)
        if "December 12, 2024" in show:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Barcelona Open 500")
        else:
            send_email("Tickets for Barcelona Open 500", "Game is on", URL_BARCA)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Barcelona Open 500")
    except Exception as e:
        send_email("Tickets for Barcelona Open 500", e, URL_BARCA)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Barcelona Open 500 or site changed")
        print("Error in check_official_site:", e)


def check_cinch():
    print("\n")
    try:
        response = requests.get(URL_CINCH, headers=HEADERS)
        soup = BeautifulSoup(response.content, 'html.parser')
        # print(soup.prettify())

        show = soup.find('div', class_='ewa-rteLine').text.strip()
        print(show)
        if "The ticket sale dates for the 2025 grass court season have not yet been confirmed." in show:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for London Cinch 500")
        else:
            send_email("Tickets for London Cinch 500", "Game is on", URL_CINCH)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                  "Tickets are now available for London Cinch 500")
    except Exception as e:
        # send_email("Tickets for London Cinch 500", e , URL_CINCH)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              "Tickets are now available for London Cinch 500 or site changed")
        print("Error in check_official_site:", e)


def check_uefa(url):
    print("\n")
    try:
        response = httpx.get(url, headers=HEADERS)
        soup = BeautifulSoup(response.content, 'html.parser')
        # print(soup.prettify())

        show = soup.find('div', class_='padding intro').text.strip()
        print(show)
        if "We plan to start the Official Hospitality sales in November" and "Champions League" in show:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Champions League")

        elif "We plan to start the Official Hospitality sales in November" and "Europa League" in show:
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Europa League")

        else:
            send_email("Tickets for Uefa Competitions", "Game is on", url)
            print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Uefa Competitions")

    except Exception as e:
        send_email("Tickets for Uefa Competitions", e, url)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Uefa Competitions or site changed")
        print("Error in check_official_site:", e)


def check_us_open():
    print("\n")
    title, description = None, None
    r = req.get(US_OPEN, impersonate="chrome124")
    soup = BeautifulSoup(r.content.decode('utf-8', 'ignore'), 'xml')

    data_reference = soup.find('data', {'reference': '_41e4b950_e2bb'})
    if data_reference:
        title = data_reference.find('title').get_text(strip=True)
        description = data_reference.find('description').get_text(strip=True)
    else:
        print("The required data reference was not found.")

    if "Thank you for attending the 2024 US Open!" in title and "We look forward to seeing you next year at the 2025 US Open! Sign up to become a US Open Insider or subscribe to text alerts to be notified when tickets go on sale." in description:
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for US OPEN 2025")
    else:
        send_email("Tickets for US OPEN 2025", "Game is on", US_OPEN)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for US OPEN 2025")


def cincinnati():
    print("\n")
    r = req.get(CINCINNATI, impersonate="chrome124")
    soup = BeautifulSoup(r.content.decode('utf-8', 'ignore'), 'html.parser')
    # print(soup.prettify())

    show = soup.find('div', class_='wrapper full-width-container').text.strip()

    normalized_show = ' '.join(show.split())

    expected_text = ("Single Session tickets for the 2025 Cincinnati Open will go on sale April 11, 2025. "
                     "For 24-hour early access to Single Session tickets, register for the 2025 ticket pre-sale. "
                     "Create your own Cincinnati Open experience by choosing the day, time, price and players that are the best fit for you. "
                     "Center Court Single Session tickets give you a reserved seat at Center Court, full access to the tournament grounds, including every match and practice court, a public parking lot, the opportunity to enjoy a variety of on-site entertainment, and the ability to explore our many dining and drink options. "
                     "Single Session tickets to P&G Grandstand Court grant you the same amenities except for access to enter Center Court. "
                     "To explore premium seating options, click here.")

    # Check if the normalized show matches the expected text
    if normalized_show == expected_text:
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "No available tickets for Cincinnati 2025")
    else:
        send_email("Tickets for Cincinnati 2025", "Game is on", CINCINNATI)
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "Tickets are now available for Cincinnati 2025")


def main():
    display = Display(visible=0, size=(1280, 720))
    display.start()

    try:

        print("\n")
        options = uc.ChromeOptions()
        options.add_argument('--blink-settings=imagesEnabled=false')
        options.binary_location = "/usr/bin/google-chrome"
        driver = uc.Chrome(driver_executable_path=driver_path, options=options)

        # CONCERTS
        # check_Adele_official_site(driver)
        # check_dua_of()

        check_coldplay_official_site()
        check_taylor_official_site(driver)
        check_taylor_ticketmaster()
        check_coldplay_ticketmaster()
        check_adele_ticketmaster()
        check_dua_ticketmaster()
        check_carpenter_ticketmaster()
        check_rodrigo_ticketmaster()
        check_eilish_ticketmaster()
        check_beyonce_ticketmaster()
        check_cabello_ticketmaster()
        check_metallica_ticketmaster()

        # TENNIS
        # check_AO_official_site()
        # check_AO_ticketmaster()
        # check_Cinch()
        # check_wimbledon(driver)
        # check_barcelona() # ΘΕΛΕΙ ΑΛΛΑΓΗ
        check_us_open()
        cincinnati()

        # FOOTBALL - BASKETBALL
        if datetime.datetime.now().day in [24, 25]:
            check_final4()
        for url in URL_UEFA:
            if url == 'https://uelf.hospitality.uefa.com/':
                continue
            check_uefa(url)

        if datetime.datetime.now().day in [5, 10, 15, 20, 25, 30] and datetime.datetime.now().hour in [8]:
            send_email("Check", "OK", "Καλά πάει το Scalping")

    except Exception as e:
        error_message = traceback.format_exc()
        send_email("Error", error_message, "No url")
        print("Error in main:", e, "\n")
    finally:
        # if driver is not None:
        #     driver.quit()
        display.stop()


if __name__ == "__main__":
    main()
